#!/usr/bin/env python3
"""
Export content/program-data.js → Excel (.xlsx) for editing in Excel / Google Sheets.
Layout matches tools/import_xls.py so you can re-import after editing.

Usage:
  pip install openpyxl
  python3 tools/export_program_to_xlsx.py
  python3 tools/export_program_to_xlsx.py /path/to/custom-output.xlsx
"""

import json
import re
import sys
from pathlib import Path

ROOT = Path(__file__).resolve().parent.parent
PROGRAM_JS = ROOT / "content" / "program-data.js"
DEFAULT_OUT = ROOT / "Tian-Hens-program-FOR-EDITING.xlsx"


def load_program() -> dict:
    text = PROGRAM_JS.read_text(encoding="utf-8")
    m = re.search(r"window\.PROGRAM\s*=\s*(\{.*\})\s*;", text, re.DOTALL)
    if not m:
        raise SystemExit(f"Could not parse window.PROGRAM from {PROGRAM_JS}")
    return json.loads(m.group(1))


def join_place(place_name: str, place_city: str) -> str:
    pn = (place_name or "").strip()
    pc = (place_city or "").strip()
    if pn and pc:
        return f"{pn}, {pc}"
    return pn or pc


def map_link_cell(map_url: str) -> str:
    u = (map_url or "").strip()
    if u.startswith("http"):
        return u
    return "Open in Google Maps"


def write_sheet(ws, program: dict) -> None:
    headers = ["Time", "Activity", "Place", "Google Maps link", "Notes"]
    r = 1

    for day in program.get("days") or []:
        ws.cell(r, 1, day.get("dateLabel") or "")
        r += 1
        for h, ci in zip(headers, range(1, 6)):
            ws.cell(r, ci, h)
        r += 1
        for row in day.get("rows") or []:
            ws.cell(r, 1, row.get("time") or "")
            ws.cell(r, 2, row.get("activity") or "")
            ws.cell(r, 3, join_place(row.get("placeName"), row.get("placeCity")))
            ws.cell(r, 4, map_link_cell(row.get("mapUrl") or ""))
            ws.cell(r, 5, row.get("notes") or "")
            r += 1
        r += 1

    pb = program.get("planB") or {}
    intro = (pb.get("intro") or "").strip() or "Replace anything in the program by these."
    ws.cell(r, 1, "Plan B")
    ws.cell(r, 5, intro)
    r += 1
    for h, ci in zip(headers, range(1, 6)):
        ws.cell(r, ci, h)
    r += 1
    for row in pb.get("rows") or []:
        ws.cell(r, 1, row.get("time") or "")
        ws.cell(r, 2, row.get("activity") or "")
        ws.cell(r, 3, row.get("place") or "")
        ws.cell(r, 4, map_link_cell(row.get("mapUrl") or ""))
        ws.cell(r, 5, row.get("notes") or "")
        r += 1
    r += 1

    oa = program.get("otherActivities") or {}
    title = (oa.get("title") or "Other potential activities").strip()
    ws.cell(r, 1, title)
    r += 1
    for row in oa.get("rows") or []:
        ws.cell(r, 2, row.get("activity") or "")
        ws.cell(r, 3, row.get("place") or "")
        ws.cell(r, 4, map_link_cell(row.get("mapUrl") or ""))
        ws.cell(r, 5, row.get("notes") or "")
        r += 1

    from openpyxl.utils import get_column_letter

    for col in range(1, 6):
        ws.column_dimensions[get_column_letter(col)].width = min(52, 12 + col * 4)


def main():
    out = Path(sys.argv[1]) if len(sys.argv) > 1 else DEFAULT_OUT
    try:
        from openpyxl import Workbook
    except ImportError:
        print("Install: pip install openpyxl", file=sys.stderr)
        raise SystemExit(1)

    program = load_program()
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    write_sheet(ws, program)
    out.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out)
    print("Wrote", out)
    print("Edit in Excel, then run: python3 tools/import_xls.py", out)


if __name__ == "__main__":
    main()
