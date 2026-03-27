#!/usr/bin/env python3
"""Import Tian's Hens party planning .xls or .xlsx -> ../content/program-data.js

Requires: pip install xlrd openpyxl
"""

import json
import re
import sys
from pathlib import Path
from urllib.parse import quote

import xlrd

ROOT = Path(__file__).resolve().parent.parent
OUT = ROOT / "content" / "program-data.js"


def normalize_name(s: str) -> str:
    return (
        s.replace("Flo’s", "Flow’s")
        .replace("Flo's", "Flow's")
        .replace(", Flo,", ", Flow,")
    )


def maps_url(place: str, link_cell: str) -> str:
    link_cell = (link_cell or "").strip()
    if link_cell.startswith("http"):
        return link_cell
    place = normalize_name((place or "").strip())
    q = re.sub(r"\s+", " ", place).strip() or "Madeira"
    return "https://www.google.com/maps/search/?api=1&query=" + quote(q)


def split_place(place: str):
    place = (place or "").strip()
    if not place:
        return "TBC", ""
    if "," in place:
        parts = [p.strip() for p in place.rsplit(",", 1)]
        if len(parts) == 2 and len(parts[1]) < 120:
            return parts[0], parts[1]
    return place, ""


def is_section_header(t0: str) -> bool:
    t0 = t0.strip()
    if not t0 or t0.lower() == "time":
        return False
    if t0.startswith("Plan B") or t0.startswith("Other potential"):
        return False
    if "—" in t0 and (
        re.search(r"(Paris|Madeira|Beauvais)", t0)
        or re.match(r"^(Mon|Tue|Wed|Thu|Fri|Sat|Sun)", t0)
    ):
        return True
    if re.match(
        r"^(Monday|Tuesday|Wednesday|Thursday|Friday|Saturday|Sunday)\s", t0, re.I
    ):
        return True
    if re.match(r"^(Mon|Tue|Wed|Thu|Fri|Sat|Sun)\s", t0, re.I) and (
        "Mar" in t0 or "Apr" in t0
    ):
        return True
    return False


def format_cell_value(book, sh, ri: int, ci: int) -> str:
    """Excel stores times as date serials; xlrd often uses XL_CELL_DATE (not NUMBER)."""
    c = sh.cell(ri, ci)
    v = c.value
    if c.ctype == xlrd.XL_CELL_EMPTY:
        return ""
    if c.ctype in (xlrd.XL_CELL_NUMBER, xlrd.XL_CELL_DATE) and isinstance(v, float):
        if 0 < v < 1:
            secs = int(round(v * 86400))
            h, m = secs // 3600, (secs % 3600) // 60
            return f"{h:02d}:{m:02d}"
        try:
            t = xlrd.xldate.xldate_as_tuple(v, book.datemode)
            if t[0] == 0 and t[1] == 0 and t[2] == 0:
                return f"{t[3]:02d}:{t[4]:02d}"
        except Exception:
            pass
        if v == int(v):
            return str(int(v))
    if isinstance(v, float) and 0 < v < 1:
        secs = int(round(v * 86400))
        h, m = secs // 3600, (secs % 3600) // 60
        return f"{h:02d}:{m:02d}"
    if v is None:
        return ""
    return str(v).strip()


def _cell_to_str_xlsx(v) -> str:
    """openpyxl cell value → string (times as HH:MM)."""
    if v is None:
        return ""
    if isinstance(v, bool):
        return str(v)
    from datetime import datetime, time

    if isinstance(v, datetime):
        return f"{v.hour:02d}:{v.minute:02d}"
    if isinstance(v, time):
        return f"{v.hour:02d}:{v.minute:02d}"
    if isinstance(v, (int, float)):
        fv = float(v)
        if 0 < fv < 1:
            secs = int(round(fv * 86400))
            h, m = secs // 3600, (secs % 3600) // 60
            return f"{h:02d}:{m:02d}"
        if fv == int(fv):
            return str(int(fv))
        return str(v).strip()
    return str(v).strip()


def load_sheet_rows(path: Path) -> list:
    """All rows as 7 columns of strings; col 0 time-normalized when applicable."""
    path = Path(path)
    if path.suffix.lower() == ".xlsx":
        from openpyxl import load_workbook

        wb = load_workbook(path, read_only=True, data_only=True)
        ws = wb.active
        rows_out = []
        for row in ws.iter_rows(values_only=True):
            vals = list(row) if row else []
            vals = vals[:7]
            while len(vals) < 7:
                vals.append(None)
            row_strs = [_cell_to_str_xlsx(c) for c in vals]
            if should_normalize_time_cell(row_strs[0]):
                row_strs[0] = normalize_time_string(row_strs[0])
            rows_out.append(row_strs)
        wb.close()
        return rows_out

    book = xlrd.open_workbook(str(path))
    sh = book.sheet_by_index(0)
    rows_out = []
    for ri in range(sh.nrows):
        out = []
        for ci in range(min(7, sh.ncols)):
            v = format_cell_value(book, sh, ri, ci)
            if ci == 0 and should_normalize_time_cell(v):
                v = normalize_time_string(v)
            out.append(v)
        while len(out) < 7:
            out.append("")
        rows_out.append(out)
    return rows_out


def _pad_hhmm(h: int, m: int) -> str:
    h = int(h) % 24
    m = int(m) % 60
    return f"{h:02d}:{m:02d}"


def _am_pm_to_24(h: str, ap: str) -> int:
    h = int(h)
    ap = ap.lower()
    if ap == "am":
        return 0 if h == 12 else h
    return 12 if h == 12 else h + 12


def _normalize_time_segment(seg: str) -> str:
    seg = seg.strip()
    if not seg:
        return seg
    m = re.match(r"^(\d{1,2}):(\d{2})\s*([ap]m)\s*$", seg, re.I)
    if m:
        hh = _am_pm_to_24(m.group(1), m.group(3))
        return _pad_hhmm(hh, m.group(2))
    m = re.match(r"^(\d{1,2}):(\d{2})\s*$", seg)
    if m:
        return _pad_hhmm(m.group(1), m.group(2))
    m = re.match(r"^(\d{1,2})\.(\d{2})\s*$", seg)
    if m:
        return _pad_hhmm(m.group(1), m.group(2))
    m = re.match(r"^(\d{1,2})[hH](\d{2})\s*$", seg)
    if m:
        return _pad_hhmm(m.group(1), m.group(2))
    m = re.match(r"^(\d{1,2})[hH]\s*$", seg)
    if m:
        return _pad_hhmm(m.group(1), 0)
    m = re.match(r"^(\d{1,2})\s*([ap]m)\s*$", seg, re.I)
    if m:
        hh = _am_pm_to_24(m.group(1), m.group(2))
        return _pad_hhmm(hh, 0)
    return seg


def normalize_time_string(s: str) -> str:
    """Uniform 24h display: HH:MM or HH:MM–HH:MM (en dash). Keeps flight-relative / free-text rows."""
    if not s or not str(s).strip():
        return s
    s = str(s).strip()
    if re.match(r"^Anytime", s, re.I):
        return s
    if "Flight time" in s:
        return s
    parts = re.split(r"\s*[–—\-]\s*", s)
    parts = [p for p in parts if p.strip()]
    if len(parts) == 1:
        return _normalize_time_segment(parts[0])
    return "–".join(_normalize_time_segment(p.strip()) for p in parts)


def should_normalize_time_cell(s: str) -> bool:
    """Do not split date headers like 'Sat 28 — Paris' on dashes."""
    s = (s or "").strip()
    if not s:
        return False
    if re.match(r"^(Anytime|Flight time)", s, re.I):
        return True
    if re.match(
        r"^(Mon|Tue|Wed|Thu|Fri|Sat|Sun)\b|^(Monday|Tuesday|Wednesday|Thursday|Friday|Saturday|Sunday)\b",
        s,
        re.I,
    ):
        return False
    if re.match(r"^(Plan B|Other potential|Time)\b", s, re.I):
        return False
    return bool(re.match(r"^\d", s))


def sanity_fix_row_time(row: dict, date_label: str) -> None:
    """Fix a few known logic clashes after time normalization."""
    t = row.get("time") or ""
    act = (row.get("activity") or "").lower()

    # Scavenger hunt 15:30–18:30; dinner at 18:00 fell inside the hunt → move dinner after.
    if "Apr 1" in date_label and "dinner" in act and t == "18:00":
        row["time"] = "19:00"

    # Golden hour before dinner (was same instant as dinner start).
    if "Mon 30" in date_label and "golden hour" in act and t == "19:00":
        row["time"] = "18:30"

    # Return car before airport buffer window.
    if "Sat Apr 4" in date_label and "return rental car" in act and t == "14:00":
        row["time"] = "13:00"


def import_xls(path: Path) -> dict:
    all_rows = load_sheet_rows(path)

    sections = []
    current = None
    plan_b_rows = []
    plan_b_intro = "Replace anything in the program by these."
    other_title = "Other potential activities"
    other_rows = []
    in_plan_b = False
    in_other = False

    for vals in all_rows:
        while len(vals) < 7:
            vals.append("")
        t0, act, place, gmap, notes = vals[0], vals[1], vals[2], vals[3], vals[4]

        if t0.strip() == "Plan B":
            if current and current["rows"]:
                sections.append(current)
            current = None
            in_plan_b = True
            in_other = False
            if notes.strip():
                plan_b_intro = notes.strip()
            continue

        if t0.strip().startswith("Other potential activities") or (
            t0.strip() == "" and act.strip().startswith("Other potential")
        ):
            in_plan_b = False
            in_other = True
            if act.strip().startswith("Other potential"):
                other_title = act.strip() or other_title
            continue

        if in_plan_b:
            if t0.strip().lower() == "time":
                continue
            if not any([t0.strip(), act.strip(), place.strip()]):
                continue
            plan_b_rows.append(
                {
                    "time": normalize_name(t0.strip() or "—"),
                    "activity": normalize_name(act.strip() or "—"),
                    "place": normalize_name(place.strip() or "—"),
                    "mapUrl": gmap.strip()
                    if gmap.startswith("http")
                    else maps_url(place, gmap),
                    "notes": normalize_name(notes.strip()),
                }
            )
            continue

        if in_other:
            if not any([act.strip(), place.strip()]):
                continue
            other_rows.append(
                {
                    "activity": normalize_name(act.strip()),
                    "place": normalize_name(place.strip()),
                    "mapUrl": gmap.strip()
                    if gmap.startswith("http")
                    else maps_url(place, gmap),
                    "notes": normalize_name(notes.strip()),
                }
            )
            continue

        if t0.strip().lower() == "time":
            continue

        if is_section_header(t0):
            if current and current["rows"]:
                sections.append(current)
            parts = t0.split("—", 1)
            theme_guess = parts[1].strip() if len(parts) > 1 else ""
            current = {
                "dateLabel": normalize_name(t0.strip()),
                "city": theme_guess or "Madeira",
                "theme": theme_guess,
                "notes": "",
                "rows": [],
            }
            continue

        if current is None:
            continue
        if not any([t0.strip(), act.strip(), place.strip()]):
            continue

        pname, pcity = split_place(place)
        activity = (act or "").strip()
        if not activity and pname:
            activity = "Stop / visit"
        if not activity:
            activity = "—"

        row_obj = {
            "time": normalize_name(t0.strip() or "—"),
            "activity": normalize_name(activity),
            "placeName": normalize_name(pname or activity),
            "placeCity": pcity,
            "mapUrl": maps_url(place, gmap),
            "notes": normalize_name(notes.strip()),
        }
        sanity_fix_row_time(row_obj, current["dateLabel"])
        current["rows"].append(row_obj)

    if current and current["rows"]:
        sections.append(current)

    return {
        "kicker": "Cha • Flow • Tian",
        "navTitle": "Hens itinerary",
        "title": "Tian’s Hens — Paris → Beauvais → Madeira",
        "dates": "Mar 28 → Apr 4",
        "location": "Paris • Beauvais • Madeira",
        "heroTitle": "Tian’s Hens",
        "subtitle": "",
        "scheduleSubtitle": "",
        "quickFacts": [
            {"label": "Group", "value": "3 (Cha, Flow, Tian)"},
        ],
        "contacts": [],
        "packingList": [
            "Comfortable shoes (Paris cobbles + Funchal)",
            "Statement / festival outfits (rainbow, sequins, fringe)",
            "Light layer + windbreaker (Madeira microclimates)",
            "Swimsuit + quick-dry towel",
            "Hiking shoes (levada) + headlamp (tunnels)",
            "Sunscreen + sunglasses",
            "Reusable water bottle",
        ],
        "planB": {
            "intro": plan_b_intro,
            "rows": plan_b_rows,
        },
        "otherActivities": {"title": other_title, "rows": other_rows},
        "days": sections,
        "footerNote": "",
    }


def main():
    xls = Path(
        "/Users/charlotte/Library/Mobile Documents/com~apple~CloudDocs/Mes fichiers/"
        "trippy/voyages/tickets and bookings/Billets 2026/Madeira/Hens party planning/"
        "Tian's Hens party planning.xls"
    )
    if len(sys.argv) > 1:
        xls = Path(sys.argv[1])

    data = import_xls(xls)
    text = "window.PROGRAM = \n" + json.dumps(data, indent=2, ensure_ascii=False) + "\n;\n"
    OUT.write_text(text, encoding="utf-8")
    print("Wrote", OUT)
    print("Days:", len(data["days"]))
    print("Plan B rows:", len(data["planB"]["rows"]))
    print("Other activities:", len(data["otherActivities"]["rows"]))


if __name__ == "__main__":
    main()
